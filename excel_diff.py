#!/usr/bin/env python3
import argparse
import difflib
import html
import json
import locale
import os
import re
import sys

from openpyxl import load_workbook
from pathlib import Path
from xlrd import open_workbook

def get_base_dir():
    """
    Return the base directory for bundled resources:
      - When frozen (bundled), return the bundle’s temp extraction directory.
      - Otherwise, return the directory of this source file.
    Works on macOS, Windows, and Linux.
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)
    return Path(__file__).resolve().parent

APP_DIR = get_base_dir()
LOCALES_DIR = APP_DIR / "locales"
DEFAULT_LOCALE = "en"

# ---------------- i18n helpers ---------------- #

def _normalize_locale_tag(tag: str) -> str:
    if not tag:
        return ""
    s = str(tag)
    s = s.split(".", 1)[0]
    s = s.split("@", 1)[0]
    s = s.strip().replace("-", "_")
    if not s:
        return ""
    parts = s.split("_")
    if len(parts) == 1:
        return parts[0].lower()
    lang = parts[0].lower()
    region = parts[1].upper()
    return f"{lang}_{region}"

def _gather_env_locales():
    seen = set()
    out = []
    for var in ("LC_ALL", "LC_MESSAGES", "LANG"):
        val = os.environ.get(var)
        if not val:
            continue
        if isinstance(val, str):
            candidates = val.split(":") if ":" in val else [val]
        else:
            candidates = [str(val)]
        for cand in candidates:
            norm = _normalize_locale_tag(cand)
            if norm and norm not in seen:
                seen.add(norm)
                out.append(norm)
                base = norm.split("_", 1)[0]
                if base and base not in seen:
                    seen.add(base)
                    out.append(base)
    return out

def detect_locale_chain():
    prefs = []
    seen = set()
    try:
        locale.setlocale(locale.LC_CTYPE, "")
        lang_tuple = locale.getlocale()
        if lang_tuple and lang_tuple[0]:
            norm = _normalize_locale_tag(lang_tuple[0])
            if norm and norm not in seen:
                seen.add(norm); prefs.append(norm)
                base = norm.split("_", 1)[0]
                if base and base not in seen:
                    seen.add(base); prefs.append(base)
    except Exception:
        pass
    for env_loc in _gather_env_locales():
        if env_loc not in seen:
            seen.add(env_loc); prefs.append(env_loc)
    if DEFAULT_LOCALE not in seen:
        prefs.append(DEFAULT_LOCALE)
    return prefs

def load_labels():
    """
    Load and merge locale files, from base to variant.
    This ensures that specific locales (e.g., en_GB) override general ones (e.g., en).
    """
    merged = {}
    # Reverse the chain to load from base to variant (e.g., 'en' then 'en_IE')
    for cand in reversed(detect_locale_chain()):
        f = LOCALES_DIR / f"{cand}_cli.json"
        if f.exists():
            try:
                with f.open("r", encoding="utf-8") as fh:
                    data = json.load(fh)
                    merged.update(data)
            except json.JSONDecodeError as e:
                # This will catch malformed JSON and tell you exactly where the error is.
                print(f"Error: Could not parse '{f.name}'. It may contain invalid JSON. Details: {e}", file=sys.stderr)
            except Exception as e:
                # Catch other potential file reading errors.
                print(f"Error: Could not read file '{f.name}'. Details: {e}", file=sys.stderr)
    return merged

# Initialize labels 
LABELS = load_labels()
def T(key, default_text):
    return LABELS.get(key, default_text)

# ---------------- Utility functions ---------------- #

def load_excel(path):
    """Return dict {sheetname: [[cell texts]]}"""
    if path is None:
        return {}
    sheets = {}
    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            sheets[ws.title] = [
                [("" if c is None else str(c)) for c in row]
                for row in ws.iter_rows(values_only=True)
            ]
    elif path.suffix.lower() == ".xls":
        wb = open_workbook(path)
        for ws in wb.sheets():
            rows = []
            for r in range(ws.nrows):
                rows.append([
                    "" if ws.cell_value(r, c) == "" else str(ws.cell_value(r, c))
                    for c in range(ws.ncols)
                ])
            sheets[ws.name] = rows
    return sheets

def safe_int(val, cap=None):
    try:
        v = int(val)
    except:
        return 0
    if v < 0:
        return 0
    if cap is not None:
        return min(v, cap)
    return v

def _wrap_del(txt):
    # red text + red background + strikethrough; preserve whitespace
    return f'<del style="color:#b00000;background-color:#ffd6d6; text-decoration:line-through">{txt}</del>'

def _wrap_ins(txt):
    # green text + green background; preserve whitespace
    return f'<span style="color:#0b6b00;background-color:#d8f8d8">{txt}</span>'

def diff_words(a, b):
    a_words = a.split()
    b_words = b.split()
    sm = difflib.SequenceMatcher(None, a_words, b_words)
    out = []
    for op, i1, i2, j1, j2 in sm.get_opcodes():
        if op == "equal":
            out.extend(html.escape(w) for w in a_words[i1:i2])
        elif op == "delete":
            out.extend(_wrap_del(html.escape(w)) for w in a_words[i1:i2])
        elif op == "insert":
            out.extend(_wrap_ins(html.escape(w)) for w in b_words[j1:j2])
        elif op == "replace":
            out.extend(_wrap_del(html.escape(w)) for w in a_words[i1:i2])
            out.extend(_wrap_ins(html.escape(w)) for w in b_words[j1:j2])
    return " ".join(out)

def diff_chars(a, b):
    sm = difflib.SequenceMatcher(None, a, b)
    out = []
    for op, i1, i2, j1, j2 in sm.get_opcodes():
        if op == "equal":
            out.append(html.escape(a[i1:i2]))
        elif op == "delete":
            out.append(_wrap_del(html.escape(a[i1:i2])))
        elif op == "insert":
            out.append(_wrap_ins(html.escape(b[j1:j2])))
        elif op == "replace":
            out.append(_wrap_del(html.escape(a[i1:i2])))
            out.append(_wrap_ins(html.escape(b[j1:j2])))
    return "".join(out)

def similarity(a, b):
    if not a and not b:
        return 100.0
    sm = difflib.SequenceMatcher(None, a, b)
    return sm.ratio() * 100.0

def col_letter_to_index(letter):
    letter = (letter or "").upper()
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1

# ---------------- Comparison logic ---------------- #
def compare_sheets(orig_rows, mod_rows, source_idx, target_idx,
                   extra_idx, row_offset, realign, tolerate,
                   omit_identical=False):
    results = []
    processed_orig = set()
    processed_mod = set()
    max_orig = len(orig_rows)
    max_mod = len(mod_rows)

    def get(row, idx):
        if row is None or idx is None:
            return None
        if idx < len(row):
            return row[idx]
        return None

    def _emit_pair(oi, mi, o_src, m_src, o_tgt, m_tgt, o_extra, m_extra):
        # Emit raw 0-based indices; renderer converts to Excel row numbers.
        # Skip rows that are completely empty (no source and no target)
        if not (o_src or o_tgt or m_src or m_tgt):
            return
        tgt_identical = int(round(similarity(o_tgt or "", m_tgt or ""))) == 100
        src_identical = (o_src or "") == (m_src or "")
        idx_identical = (oi == mi)
        if omit_identical:
            # Strict: omit any identical targets
            if tgt_identical:
                return
        else:
            # When omit_identical is False:
            # omit rows where target, index, and source are ALL identical
            if tgt_identical and idx_identical and src_identical:
                return
        results.append((oi, mi, o_src, m_src, o_tgt, m_tgt, o_extra, m_extra))

    def _src_similarity(a, b):
        # similarity() already returns 0..100 using difflib.SequenceMatcher
        return similarity(a or "", b or "")

    # Find best matching modified-row index for given original-row index
    def find_best_match_for_oi(oi):
        o_row = orig_rows[oi]
        o_src = get(o_row, source_idx)

        best_mi = None
        best_score = -1.0
        threshold = max(0.0, 100.0 - float(tolerate or 0))

        # Pass 1: similarity-based candidates honoring tolerate
        for mi in range(max_mod):
            if mi in processed_mod:
                continue
            m_row = mod_rows[mi]
            m_src = get(m_row, source_idx)
            sim = _src_similarity(o_src, m_src)
            if sim >= threshold:
                score = sim
                if (o_src or "") == (m_src or ""):
                    score += 10.0  # strong boost for exact equality
                if oi == mi:
                    score += 2.0   # small boost for same index
                if score > best_score:
                    best_score = score
                    best_mi = mi

        # Pass 2: if none passed threshold, exact source match within ±realign window
        if best_mi is None and realign and realign > 0 and (o_src is not None):
            for delta in range(1, realign + 1):
                for mi in (oi - delta, oi + delta):
                    if mi < 0 or mi >= max_mod or mi in processed_mod:
                        continue
                    m_row = mod_rows[mi]
                    m_src = get(m_row, source_idx)
                    if (o_src or "") == (m_src or ""):
                        return mi

        return best_mi

    # Main loop over original rows
    for oi in range(max_orig):
        if oi in processed_orig:
            continue

        o_row = orig_rows[oi]
        o_src = get(o_row, source_idx)
        o_tgt = get(o_row, target_idx)
        o_extra = get(o_row, extra_idx)

        best_mi = find_best_match_for_oi(oi)

        if best_mi is not None:
            processed_orig.add(oi)
            processed_mod.add(best_mi)
            m_row = mod_rows[best_mi]
            m_src = get(m_row, source_idx)
            m_tgt = get(m_row, target_idx)
            m_extra = get(m_row, extra_idx)
            # Emit raw 0-based indices; renderer will convert for display
            _emit_pair(oi, best_mi, o_src, m_src, o_tgt, m_tgt, o_extra, m_extra)
        else:
            # No match -> deletion
            _emit_pair(oi, None, o_src, None, o_tgt, None, o_extra, None)

    # Leftover modified rows -> insertions
    for mi in range(max_mod):
        if mi in processed_mod:
            continue
        m_row = mod_rows[mi]
        m_src = get(m_row, source_idx)
        m_tgt = get(m_row, target_idx)
        m_extra = get(m_row, extra_idx)
        _emit_pair(None, mi, None, m_src, None, m_tgt, None, m_extra)

    return results

# ---------------- HTML Rendering ---------------- #

def style_css():
    return T("STYLE", """
<style>
  body { font-family: system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif; }
  /* Use fixed layout so column widths from colgroup are honored */
  table { width: 100%; border-collapse: collapse; margin: 12px 0; table-layout: fixed; }
  th, td { border: 1px solid #ccc; padding: 6px 8px; vertical-align: middle; }
  /* All headers: center horizontally and middle vertically */
  th { background: #a6a6a6; text-align: center; vertical-align: middle; }
  td, th { white-space: pre-wrap; }

  /* Fixed-width columns via colgroup */
  col.line-col { width: 4em; }
  col.sim-col  { width: 5.5em; }

  /* Body cells for numeric columns: right + middle */
  td.line-col { text-align: right; vertical-align: middle; }
  td.sim-col  { text-align: right; vertical-align: middle; }

  .src-diff { border: 1px solid #c9c9c9; background: #fafafa; padding: 6px; margin-top: 6px; }
  .toc ul { margin: 0 0 6px 18px; padding: 0; }
  .toc a { text-decoration: none; }
  hr { border: none; border-top: 1px solid #888; }

  /* Zebra striping: apply only when table has class 'zebra' and only to tbody rows */
  table.zebra tbody tr:nth-child(odd) td { background-color: #ffffff; }
  table.zebra tbody tr:nth-child(even) td { background-color: #efefef; }
</style>
""")

def _same_or_both_null(a, b):
    a_empty = (a is None) or (a == "")
    b_empty = (b is None) or (b == "")
    if a_empty and b_empty:
        return True
    return (a or "") == (b or "")

def render_html(all_results, output_path, extra_header, row_offset=0):
    toc = [f"<div class='toc'><h2>{html.escape(T('TOC_TITLE', 'Table of Contents'))}</h2><ul>"]
    body = []
    for file_key, sheets in all_results.items():
        anchor_file = f"f_{re.sub(r'[^A-Za-z0-9_]+', '_', file_key)}"
        toc.append(f'<li><a href="#{anchor_file}">{html.escape(file_key)}</a><ul>')
        body.append(f'<h2 id="{anchor_file}">{html.escape(file_key)}</h2>')
        for sheetname, rows in sheets.items():
            sid = f"{anchor_file}_s_{re.sub(r'[^A-Za-z0-9_]+', '_', sheetname)}"
            toc.append(f'<li><a href="#{sid}">{html.escape(sheetname)}</a></li>')
            body.append(f'<h3 id="{sid}">{html.escape(sheetname)}</h3>')
            zebra_class = " class='zebra'" if len(rows) >= 3 else ""
            body.append(f'<table{zebra_class}>')
            # Fixed-width columns via colgroup (first and last columns)
            if extra_header:
                # Columns: Line | Extra | Source | Orig | Mod | Word diff | Char diff | Sim
                body.append(
                    "<colgroup>"
                    "<col class='line-col'>"
                    "<col>"
                    "<col>"
                    "<col>"
                    "<col>"
                    "<col>"
                    "<col>"
                    "<col class='sim-col'>"
                    "</colgroup>"
                )
            else:
                # Columns: Line | Source | Orig | Mod | Word diff | Char diff | Sim
                body.append(
                    "<colgroup>"
                    "<col class='line-col'>"
                    "<col>"
                    "<col>"
                    "<col>"
                    "<col>"
                    "<col>"
                    "<col class='sim-col'>"
                    "</colgroup>"
                )
            headers = [T("HDR_LINE", "Line")]
            if extra_header:
                headers.append(html.escape(extra_header))
            headers.extend([
                T("HDR_SOURCE", "Source"),
                T("HDR_ORIG_TGT", "Original target"),
                T("HDR_MOD_TGT", "Modified target"),
                T("HDR_TGT_WORD_DIFF", "Target diff by word"),
                T("HDR_TGT_CHAR_DIFF", "Target diff by character"),
                T("HDR_TGT_SIM", "Target similarity"),
            ])
            # Add classes to the first and last header cells for clarity
            thead_cells = []
            for idx, h in enumerate(headers):
                if idx == 0:
                    thead_cells.append(f"<th class='line-col'>{h}</th>")
                elif idx == len(headers) - 1:
                    thead_cells.append(f"<th class='sim-col'>{h}</th>")
                else:
                    thead_cells.append(f"<th>{h}</th>")
            body.append("<thead><tr>" + "".join(thead_cells) + "</tr></thead>")
            body.append("<tbody>")

            # Convert 0-based internal index to Excel row number once per side
            def _disp(idx):
                if idx is None:
                    return ""
                return str(idx + row_offset + 1)

            for oi, mi, o_src, m_src, o_tgt, m_tgt, o_extra, m_extra in rows:
                # Build the human-visible Line cell
                disp_oi = _disp(oi)
                disp_mi = _disp(mi)
                if disp_oi and disp_mi and disp_oi == disp_mi:
                    line = disp_oi
                else:
                    line = f"{disp_oi}/{disp_mi}"

                t_sim_int = int(round(similarity(o_tgt or "", m_tgt or "")))

                if extra_header:
                    if _same_or_both_null(o_extra, m_extra):
                        extra_col = html.escape((o_extra or ""))
                    else:
                        extra_col = f"{html.escape(o_extra or '')}<hr>{html.escape(m_extra or '')}"
                else:
                    extra_col = None

                if (o_src or "") == (m_src or ""):
                    src_html = html.escape(o_src or "")
                else:
                    base = f"{html.escape(o_src or '')}<hr>{html.escape(m_src or '')}"
                    src_diff = (
                        f"<strong>{html.escape(T('WORD_DIFF', 'Word diff:'))}</strong> " + diff_words(o_src or "", m_src or "") + "<br>"
                        f"<strong>{html.escape(T('CHAR_DIFF', 'Char diff:'))}</strong> </strong>" + diff_chars(o_src or "", m_src or "")
                    )
                    src_html = f"{base}<div class='src-diff'>{src_diff}</div>"

                td_word = ""
                td_char = ""
                if t_sim_int < 100:
                    td_word = diff_words(o_tgt or "", m_tgt or "")
                    td_char = diff_chars(o_tgt or "", m_tgt or "")

                row_cells = []
                row_cells.append(f"<td class='line-col'>{line}</td>")
                if extra_header:
                    row_cells.append(f"<td>{extra_col}</td>")
                row_cells.extend([
                    f"<td>{src_html}</td>",
                    f"<td>{html.escape(o_tgt or '')}</td>",
                    f"<td>{html.escape(m_tgt or '')}</td>",
                    f"<td>{td_word}</td>",
                    f"<td>{td_char}</td>",
                    f"<td class='sim-col'>{t_sim_int}%</td>",
                ])
                body.append("<tr>" + "".join(row_cells) + "</tr>")

            body.append("</tbody>")
            body.append("</table>")
        toc.append("</ul></li>")
    toc.append("</ul></div>")
    with Path(output_path).open("w", encoding="utf-8") as f:
        f.write("<html><head><meta charset='utf-8'>")
        f.write(style_css())
        f.write("</head><body>")
        f.write("\n".join(toc))
        f.write("\n".join(body))
        f.write("</body></html>")

# ---------------- Main ---------------- #

def main():
    ap = argparse.ArgumentParser(description=T("CLI_DESC", "Compare Excel files or folders and render an HTML diff report."))
    ap.add_argument("--original", required=True, help=T("ARG_HELP_ORIGINAL", "Path to original file or directory (with --dir)"))
    ap.add_argument("--modified", required=True, help=T("ARG_HELP_MODIFIED", "Path to modified file or directory (with --dir)"))
    ap.add_argument("--output", required=True, help=T("ARG_HELP_OUTPUT", "Path to output HTML file"))
    ap.add_argument("--source", required=True, help=T("ARG_HELP_SOURCE", "Column letter for source"))
    ap.add_argument("--target", required=True, help=T("ARG_HELP_TARGET", "Column letter for target"))
    ap.add_argument("--extra_column", help=T("ARG_HELP_EXTRA_COL", "Column letter for optional extra text"))
    ap.add_argument("--extra_header", default=T("ARG_DEF_EXTRA_HEADER", "Extra column"), help=T("ARG_HELP_EXTRA_HEADER", "Header for the extra column in the HTML"))
    ap.add_argument("--row-offset", type=int, default=0, help=T("ARG_HELP_ROW_OFFSET", "Header row offset (rows to skip)"))
    ap.add_argument("--realign", type=int, default=0, help=T("ARG_HELP_REALIGN", "Realign search window (rows)"))
    ap.add_argument("--tolerate", type=int, default=0, help=T("ARG_HELP_TOLERATE", "Similarity tolerance (%)"))
    ap.add_argument("--nocap", action="store_true", help=T("ARG_HELP_NOCAP", "Remove caps on realign/tolerance"))
    ap.add_argument("--dir", action="store_true", help=T("ARG_HELP_DIR", "Treat --original and --modified as directories"))
    ap.add_argument("--wspattern", help=T("ARG_HELP_WSPATTERN", "Regex applied to worksheet names (tabs)"))
    ap.add_argument("--omit_identical", action="store_true", help=T("ARG_HELP_OMIT_IDENTICAL", "Omit rows where targets are identical (100% similarity)"))
    args = ap.parse_args()

    cap_realign = None if args.nocap else 15
    cap_tolerate = 100 if args.nocap else 35
    realign = safe_int(args.realign, cap_realign)
    tolerate = safe_int(args.tolerate, cap_tolerate)

    source_idx = col_letter_to_index(args.source)
    target_idx = col_letter_to_index(args.target)
    extra_idx = col_letter_to_index(args.extra_column) if args.extra_column else None

    # Pre-compile worksheet-name pattern (if provided)
    ws_pat = None
    if args.wspattern:
        try:
            ws_pat = re.compile(args.wspattern)
        except re.error:
            ws_pat = None  # invalid regex -> behave as if no filter

    orig = Path(args.original)
    mod = Path(args.modified)
    out_path = Path(args.output)

    # ---- Validation: enforce dir/file consistency ----
    if args.dir:
        if not orig.exists() or not orig.is_dir():
            raise SystemExit(T("ERR_ORIG_DIR", "--original must be an existing directory when --dir is used: ") + str(orig))
        if not mod.exists() or not mod.is_dir():
            raise SystemExit(T("ERR_MOD_DIR", "--modified must be an existing directory when --dir is used: ") + str(mod))
    else:
        if not orig.exists() or not orig.is_file():
            raise SystemExit(T("ERR_ORIG_FILE", "--original must be an existing file when --dir is NOT used: ") + str(orig))
        if not mod.exists() or not mod.is_file():
            raise SystemExit(T("ERR_MOD_FILE", "--modified must be an existing file when --dir is NOT used: ") + str(mod))

    def list_excels(folder: Path):
        # Walk a directory and yield Excel files (no wspattern on filenames)
        for root, _, files in os.walk(folder):
            root_p = Path(root)
            for fn in files:
                if not fn.lower().endswith((".xls", ".xlsx")):
                    continue
                if fn.startswith("~$"):
                    continue  # skip Excel temp/lock files
                yield root_p / fn

    results = {}
    if args.dir:
        # Directory comparison: align by filename across both trees
        orig_files = {f.name: f for f in list_excels(orig)}
        mod_files = {f.name: f for f in list_excels(mod)}
        all_keys = sorted(set(orig_files) | set(mod_files))
        for k in all_keys:
            o = orig_files.get(k)
            m = mod_files.get(k)
            sheets_o = load_excel(o)
            sheets_m = load_excel(m)
            file_key = k
            results[file_key] = {}
            all_sheets = set(sheets_o) | set(sheets_m)
            # Filter by worksheet (tab) name using --wspattern, if provided
            if ws_pat is not None:
                all_sheets = {sh for sh in all_sheets if ws_pat.search(sh)}
            for sh in sorted(all_sheets):
                # Strictly enforce row-offset by slicing before comparison
                o_rows_all = sheets_o.get(sh, [])
                m_rows_all = sheets_m.get(sh, [])
                o_rows = o_rows_all[args.row_offset:] if args.row_offset > 0 else o_rows_all
                m_rows = m_rows_all[args.row_offset:] if args.row_offset > 0 else m_rows_all
                rows = compare_sheets(
                    o_rows,
                    m_rows,
                    source_idx, target_idx,
                    extra_idx,
                    args.row_offset,
                    realign,
                    tolerate,
                    omit_identical=args.omit_identical
                )
                results[file_key][sh] = rows
    else:
        # Single file comparison: do not treat as directories
        sheets_o = load_excel(orig)
        sheets_m = load_excel(mod)
        file_key = f"{orig.name}/{mod.name}" if orig.name != mod.name else orig.name
        results[file_key] = {}
        all_sheets = set(sheets_o) | set(sheets_m)
        # Filter by worksheet (tab) name using --wspattern, if provided
        if ws_pat is not None:
            all_sheets = {sh for sh in all_sheets if ws_pat.search(sh)}
        for sh in sorted(all_sheets):
            # Strictly enforce row-offset by slicing before comparison
            o_rows_all = sheets_o.get(sh, [])
            m_rows_all = sheets_m.get(sh, [])
            o_rows = o_rows_all[args.row_offset:] if args.row_offset > 0 else o_rows_all
            m_rows = m_rows_all[args.row_offset:] if args.row_offset > 0 else m_rows_all
            rows = compare_sheets(
                o_rows,
                m_rows,
                source_idx, target_idx,
                extra_idx,
                args.row_offset,
                realign,
                tolerate,
                omit_identical=args.omit_identical
            )
            results[file_key][sh] = rows

    render_html(results, out_path, args.extra_header if args.extra_column else None, row_offset=args.row_offset)

if __name__ == "__main__":
    main()
