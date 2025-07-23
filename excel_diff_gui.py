import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import re
import os
import html
from difflib import ndiff
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# === Diff and HTML generation logic ===

def get_diff_html(orig, mod, by_word=True):
    orig = orig or ""
    mod = mod or ""
    tokenize = (lambda x: x.split()) if by_word else list
    diff = ndiff(tokenize(orig), tokenize(mod))
    result = []
    for d in diff:
        op, text = d[0], html.escape(d[2:])
        if op == ' ':
            result.append(text)
        elif op == '+':
            result.append(f'<span style="color:green;">{text}</span>')
        elif op == '-':
            result.append(f'<span style="color:red;text-decoration:line-through;">{text}</span>')
    return (' ' if by_word else '').join(result)

def process_sheet(ws_orig, ws_mod, source_col, target_col, row_offset, by_word):
    max_row = max(ws_orig.max_row, ws_mod.max_row)
    rows = []
    for r in range(row_offset + 1, max_row + 1):
        source = ws_orig.cell(row=r, column=source_col).value or ""
        orig_val = ws_orig.cell(r, column=target_col).value or ""
        mod_val  = ws_mod.cell(r, column=target_col).value or ""
        if orig_val != mod_val:
            diff_view = get_diff_html(orig_val, mod_val, by_word)
            rows.append((r, source, orig_val, mod_val, diff_view))
    return rows

def process_sheet_as_deletion(ws, source_col, target_col, row_offset, by_word):
    rows = []
    for r in range(row_offset + 1, ws.max_row + 1):
        source = ws.cell(r, column=source_col).value or ""
        orig_val = ws.cell(r, column=target_col).value or ""
        mod_val = ""
        if orig_val:
            diff_view = get_diff_html(orig_val, mod_val, by_word)
            rows.append((r, source, orig_val, mod_val, diff_view))
    return rows

def process_sheet_as_insertion(ws, source_col, target_col, row_offset, by_word):
    rows = []
    for r in range(row_offset + 1, ws.max_row + 1):
        source = ws.cell(r, column=source_col).value or ""
        orig_val = ""
        mod_val = ws.cell(r, column=target_col).value or ""
        if mod_val:
            diff_view = get_diff_html(orig_val, mod_val, by_word)
            rows.append((r, source, orig_val, mod_val, diff_view))
    return rows

def compare_workbook_pair(file_label, original_path, modified_path, wspattern, source_col, target_col, row_offset, by_word):
    try:
        wb_o = load_workbook(original_path, data_only=True)
        wb_m = load_workbook(modified_path, data_only=True)
    except Exception:
        return None

    sheet_results = []
    for name in wb_o.sheetnames:
        if not wspattern.match(name) or name not in wb_m.sheetnames:
            continue
        rows = process_sheet(wb_o[name], wb_m[name], source_col, target_col, row_offset, by_word)
        if rows:
            sheet_results.append((name, rows))
    return (file_label, sheet_results) if sheet_results else None

def write_html_report(filename, all_results):
    with open(filename, "w", encoding="utf-8") as f:
        f.write("<!DOCTYPE html><html><head><meta charset='utf-8'>\n")
        f.write("<title>Excel Diff Report</title>\n")
        f.write("""
        <style>
            body { font-family: Arial, sans-serif; padding: 20px; }
            table { border-collapse: collapse; margin-bottom: 2em; width: 100%; }
            th, td { border: 1px solid #ccc; padding: 8px; vertical-align: top; }
            th { background: #f0f0f0; }
            a { color: #0077cc; }
            .toc ul { list-style-type: none; padding-left: 1em; }
            .toc li { margin-bottom: 4px; }
            .toc .file { font-weight: bold; margin-top: 1em; }
            pre { white-space: pre-wrap; }
        </style>
        </head><body>
        <h1>Excel Diff Report</h1>
        <div class="toc"><h2>Table of Contents</h2>
        <ul>
        """)

        for file_label, sheets in all_results:
            anchor_file = f"file--{html.escape(file_label)}"
            f.write(f"<li class='file'><a href='#{anchor_file}'>{html.escape(file_label)}</a></li>\n")
            f.write("<ul>\n")
            for sheet_name, _ in sheets:
                sheet_anchor = f"{anchor_file}--{sheet_name}"
                f.write(f"<li><a href='#{sheet_anchor}'>{html.escape(sheet_name)}</a></li>\n")
            f.write("</ul>\n")
        f.write('</ul></div>\n')

        for file_label, sheets in all_results:
            file_anchor = f"file--{html.escape(file_label)}"
            f.write(f"<h2 id='{file_anchor}'>{html.escape(file_label)}</h2>\n")
            for sheet_name, rows in sheets:
                sheet_anchor = f"{file_anchor}--{sheet_name}"
                f.write(f"<h3 id='{sheet_anchor}'>{html.escape(sheet_name)}</h3>\n")
                f.write("<table>\n")
                f.write("<tr><th>Line</th><th>Source</th><th>Original Target</th><th>Modified Target</th><th>Diff</th></tr>\n")
                for line, src, orig, mod, diff_html in rows:
                    f.write("<tr>\n")
                    f.write(f"<td>{line}</td>")
                    f.write(f"<td>{html.escape(str(src))}</td>")
                    f.write(f"<td>{html.escape(str(orig))}</td>")
                    f.write(f"<td>{html.escape(str(mod))}</td>")
                    f.write(f"<td><pre>{diff_html}</pre></td>")
                    f.write("</tr>\n")
                f.write("</table>\n")

        f.write("</body></html>\n")

# === Tkinter GUI Application ===

class DiffApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Diff GUI")
        self.geometry("")
        self.update_idletasks()
        self.minsize(self.winfo_reqwidth(), self.winfo_reqheight())

        self.original_path = tk.StringVar()
        self.modified_path = tk.StringVar()
        self.dir_mode = tk.BooleanVar(value=True)
        self.wspattern = tk.StringVar(value=".*")
        self.source_col = tk.StringVar(value="B")
        self.target_col = tk.StringVar(value="C")
        self.row_offset = tk.IntVar(value=1)
        self.diff_mode = tk.StringVar()

        home = os.path.expanduser("~")
        self.last_dir = tk.StringVar(value=home)
        self.last_output_dir = tk.StringVar(value=home)
        self.output_path = tk.StringVar(value=os.path.join(home, "diff_report.html"))

        self.create_widgets()

    def exit_app(self):
        self.destroy()
        self.quit()

    def create_widgets(self):
        pad = {'padx': 5, 'pady': 5}
        tk.Checkbutton(self, text="Directory mode", variable=self.dir_mode).grid(row=0, column=1, sticky="w", **pad)

        tk.Label(self, text="Original file/dir:").grid(row=1, column=0, sticky="e", **pad)
        tk.Entry(self, textvariable=self.original_path, width=40).grid(row=1, column=1, **pad)
        tk.Button(self, text="Browse", command=self.browse_original).grid(row=1, column=2, **pad)

        tk.Label(self, text="Modified file/dir:").grid(row=2, column=0, sticky="e", **pad)
        tk.Entry(self, textvariable=self.modified_path, width=40).grid(row=2, column=1, **pad)
        tk.Button(self, text="Browse", command=self.browse_modified).grid(row=2, column=2, **pad)

        tk.Label(self, text="Worksheet name pattern:").grid(row=3, column=0, sticky="e", **pad)
        tk.Entry(self, textvariable=self.wspattern, width=40).grid(row=3, column=1, columnspan=2, sticky="w", **pad)

        tk.Label(self, text="Source column letter:").grid(row=4, column=0, sticky="e", **pad)
        tk.Entry(self, textvariable=self.source_col, width=5).grid(row=4, column=1, sticky="w", **pad)

        tk.Label(self, text="Target column letter:").grid(row=5, column=0, sticky="e", **pad)
        tk.Entry(self, textvariable=self.target_col, width=5).grid(row=5, column=1, sticky="w", **pad)

        tk.Label(self, text="Rows to skip (offset):").grid(row=6, column=0, sticky="e", **pad)
        tk.Spinbox(self, from_=0, to=1000, textvariable=self.row_offset, width=5).grid(row=6, column=1, sticky="w", **pad)

        tk.Label(self, text="Diff mode:").grid(row=7, column=0, sticky="e", **pad)
        self.diff_mode_dropdown = ttk.Combobox(self, textvariable=self.diff_mode, state="readonly", width=18)
        self.diff_mode_dropdown['values'] = ("Word", "Character")
        self.diff_mode_dropdown.grid(row=7, column=1, columnspan=2, sticky="w", **pad)
        self.diff_mode_dropdown.set("Word")

        tk.Label(self, text="Output HTML file:").grid(row=8, column=0, sticky="e", **pad)
        tk.Entry(self, textvariable=self.output_path, width=40).grid(row=8, column=1, **pad)
        tk.Button(self, text="Browse", command=self.browse_output).grid(row=8, column=2, **pad)

        tk.Button(self, text="Run Diff", command=self.run_diff).grid(row=9, column=2, sticky="w", **pad)
        tk.Button(self, text="Exit", command=self.exit_app).grid(row=11, column=2, sticky="e", **pad)

        self.progress_label = tk.Label(self, text="")
        self.progress_label.grid(row=10, column=0, columnspan=3)

    def browse_original(self):
        if self.dir_mode.get():
            path = filedialog.askdirectory(initialdir=self.last_dir.get())
        else:
            path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")], initialdir=self.last_dir.get())
        if path:
            self.original_path.set(path)
            self.last_dir.set(os.path.dirname(path) if os.path.isfile(path) else path)

    def browse_modified(self):
        if self.dir_mode.get():
            path = filedialog.askdirectory(initialdir=self.last_dir.get())
        else:
            path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")], initialdir=self.last_dir.get())
        if path:
            self.modified_path.set(path)
            self.last_dir.set(os.path.dirname(path) if os.path.isfile(path) else path)

    def browse_output(self):
        orig_name = os.path.basename(self.original_path.get())
        base_name = os.path.splitext(orig_name)[0] if orig_name else "diff"
        suggested_name = f"{base_name}_diff.html"
        path = filedialog.asksaveasfilename(
            defaultextension=".html",
            initialfile=suggested_name,
            filetypes=[("HTML files", "*.html")],
            initialdir=self.last_output_dir.get()
        )
        if path:
            self.output_path.set(path)
            self.last_output_dir.set(os.path.dirname(path))

    def run_diff(self):
        if not self.original_path.get() or not self.modified_path.get():
            messagebox.showerror("Error", "Please select original and modified paths.")
            return
        self.progress_label.config(text="Running diff... please wait.")
        self.update_idletasks()
        threading.Thread(target=self.run_diff_thread, daemon=True).start()

    def run_diff_thread(self):
        try:
            mode_label = self.diff_mode_dropdown.get().lower()
            by_word = (mode_label == "word")
            wspattern = re.compile(self.wspattern.get())
            source_col = column_index_from_string(self.source_col.get().strip().upper())
            target_col = column_index_from_string(self.target_col.get().strip().upper())
            row_offset = self.row_offset.get()
            output_file = self.output_path.get()
            results = []

            if self.dir_mode.get():
                orig_dir = self.original_path.get()
                mod_dir = self.modified_path.get()
                orig_files = {
                    f for f in os.listdir(orig_dir)
                    if f.lower().endswith((".xls", ".xlsx")) and os.path.isfile(os.path.join(orig_dir, f))
                }
                mod_files = {
                    f for f in os.listdir(mod_dir)
                    if f.lower().endswith((".xls", ".xlsx")) and os.path.isfile(os.path.join(mod_dir, f))
                }
                all_files = sorted(orig_files | mod_files)
                for f in all_files:
                    orig_file = os.path.join(orig_dir, f)
                    mod_file = os.path.join(mod_dir, f)
                    if f in orig_files and f in mod_files:
                        r = compare_workbook_pair(f, orig_file, mod_file, wspattern, source_col, target_col, row_offset, by_word)
                        if r:
                            results.append(r)
                    elif f in orig_files:
                        wb = load_workbook(orig_file, data_only=True)
                        sheet_results = []
                        for name in wb.sheetnames:
                            if wspattern.match(name):
                                rows = process_sheet_as_deletion(wb[name], source_col, target_col, row_offset, by_word)
                                if rows:
                                    sheet_results.append((name, rows))
                        if sheet_results:
                            results.append((f"{f} (deleted)", sheet_results))
                    elif f in mod_files:
                        wb = load_workbook(mod_file, data_only=True)
                        sheet_results = []
                        for name in wb.sheetnames:
                            if wspattern.match(name):
                                rows = process_sheet_as_insertion(wb[name], source_col, target_col, row_offset, by_word)
                                if rows:
                                    sheet_results.append((name, rows))
                        if sheet_results:
                            results.append((f"{f} (added)", sheet_results))
            else:
                label = os.path.basename(self.original_path.get())
                result = compare_workbook_pair(label, self.original_path.get(), self.modified_path.get(),
                                               wspattern, source_col, target_col, row_offset, by_word)
                if result:
                    results.append(result)

            if results:
                write_html_report(output_file, results)
                self.progress_label.config(text=f"✅ Report saved: {output_file}")
                messagebox.showinfo("Success", f"Report generated:\n{output_file}")
            else:
                self.progress_label.config(text="")
                messagebox.showinfo("No Differences", "No differences found or matching files.")
        except Exception as e:
            self.progress_label.config(text="")
            messagebox.showerror("Error", f"An error occurred:\n{e}")

if __name__ == "__main__":
    app = DiffApp()
    app.mainloop()
